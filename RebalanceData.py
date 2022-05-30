from itertools import chain
import json
from multiprocessing import pool
from pyparsing import empty
import requests
import pandas as pd
import schedule
import time
import openpyxl
import xlsxwriter
from openpyxl import load_workbook
from os.path import exists

def write_to_excel(tokenAddress):
    request = requests.get("https://api.tracer.finance/poolsv2/upkeeps?network=42161&poolAddress=" + tokenAddress)
    request_text = request.text
    JSON = json.loads(request_text)
    excel_file = 'data/' + tokenAddress + '.xlsx'

    filterJSON = {
        'timestamp': time.strftime('%Y-%m-%d T %H:%M:%S %Z',time.gmtime(JSON['rows'][0]['blockTimestamp'])),
        'index': int(JSON['rows'][0]['endPrice'])/1e18,
        'ltoken': int(JSON['rows'][0]['longTokenPrice'])/1e06,
        'stoken': int(JSON['rows'][0]['shortTokenPrice'])/1e06,
        
    }

    df = pd.DataFrame(filterJSON, index=[filterJSON['timestamp']]
                    ,columns=['timestamp','index', 'ltoken', 'stoken'])
    
    if exists(excel_file) == False:
        workbook = xlsxwriter.Workbook(excel_file)
        worksheet = workbook.add_worksheet("Time Series Data")
        worksheet.write('A1', 'timestamp') 
        worksheet.write('B1', 'index') 
        worksheet.write('C1', 'ltoken') 
        worksheet.write('D1', 'stoken') 
        workbook.close()
        # Adding twice as theres a bug reading in the excel file with one value logged.
        add_value(excel_file ,df)
        add_value(excel_file ,df)

    elif check_exists(excel_file, df) == False:
        add_value(excel_file ,df)


def check_exists(excel_file ,df):
    existingDF = pd.read_excel(excel_file, engine='openpyxl' ,sheet_name='Time Series Data')
    if existingDF.empty == False and existingDF['timestamp'][len(df)-1] == df['timestamp'][0]:
        print('value already added')
        return True
    else:
        return False

def add_value(excel_file ,df):
    rows = df.values.tolist()
    workbook = load_workbook(excel_file)
    sheet = workbook['Time Series Data']
    for row in rows:
        sheet.append(row)
    workbook.save(excel_file)
    workbook.close()

def get_pool_list():
    pools_file = 'data/pools.xlsx'
    if exists(pools_file) == False:
        workbook = xlsxwriter.Workbook(pools_file)
        worksheet = workbook.add_worksheet("pools")
        worksheet.write('A1', 'Address')  
        workbook.close()

    request = requests.get("https://api.tracer.finance/poolsv2/poolList?network=42161&list=verified")
    request_text = request.text
    JSON = json.loads(request_text)
    list = []
    for pool in JSON['pools']:
        list.append(pool['address'])

    workbook = load_workbook(pools_file)
    idx= workbook.sheetnames.index('pools')
    ws = workbook.get_sheet_by_name('pools')
    workbook.remove(ws)
    

    workbook.create_sheet('pools', idx)
    sheet = workbook['pools']
    sheet.insert_rows(1)

    # Create the pandas DataFrame
    df = pd.DataFrame(list, columns = ['Address'])
    rows = df.values.tolist()
    for row in rows:
        sheet.append(row)
    workbook.save(pools_file)
    workbook.close()

    
def schedule_pools():
    existingPools = pd.read_excel('data/pools.xlsx', engine='openpyxl' ,sheet_name='pools')
    existingPools = existingPools.dropna()
    rows = list(chain.from_iterable(existingPools.values.tolist()))
    for row in rows:
        write_to_excel(row)
        time.sleep(10)

        

if __name__ == "__main__":
    schedule.every(1).day.do(get_pool_list)
    schedule.every(5).minutes.do(schedule_pools)
    while True:
        schedule.run_pending()
        


